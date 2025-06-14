from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, send_file
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from decimal import Decimal
from models import db, Item, SaidaMaterial, SaidaItem, Usuario, UnidadeLocal, RequisicaoMaterial
from sqlalchemy import desc, and_, or_, func
import pandas as pd
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.orm import aliased

saida_bp = Blueprint('saida_bp', __name__)

def format_currency(value):
    """Formata valor monetário para exibição"""
    if value is None:
        return "R$ 0,00"
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def parse_date(date_str):
    """Converte string de data para objeto date"""
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except:
        return None

def gerar_numero_documento():
    """Gera número sequencial do documento de saída para o ano atual"""
    ano_atual = date.today().year
    ultimo_documento = SaidaMaterial.query.filter(
        db.extract('year', SaidaMaterial.data_movimento) == ano_atual
    ).order_by(SaidaMaterial.id.desc()).first()

    if ultimo_documento and ultimo_documento.numero_documento:
        try:
            ultimo_numero = int(ultimo_documento.numero_documento.split('/')[0].replace('SAIDA', ''))
            novo_numero = ultimo_numero + 1
        except:
            novo_numero = 1
    else:
        novo_numero = 1

    return f"SAIDA{novo_numero:04d}/{ano_atual}"

@saida_bp.route('/saidas')
@login_required
def listar_saidas():
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # Filtros
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    solicitante_id = request.args.get('solicitante_id', type=int)
    item_id = request.args.get('item_id', type=int)
    status = request.args.get('status', 'EFETIVADA')  # Por padrão, mostra apenas saídas efetivadas
    filtro = request.args.get('filtro', 'data')
    busca = request.args.get('busca', '')
    
    # Query base com joins necessários usando aliases
    usuario_responsavel = aliased(Usuario)
    usuario_solicitante = aliased(Usuario)
    
    query = SaidaMaterial.query\
        .outerjoin(usuario_responsavel, SaidaMaterial.usuario_id == usuario_responsavel.id)\
        .outerjoin(usuario_solicitante, SaidaMaterial.solicitante_id == usuario_solicitante.id)\
        .outerjoin(UnidadeLocal, usuario_solicitante.unidade_local_id == UnidadeLocal.id)\
        .filter(SaidaMaterial.estornada == False)
    
    # Aplicar filtros de busca
    if busca:
        if filtro == 'data':
            try:
                data_busca = datetime.strptime(busca, '%d/%m/%Y').date()
                query = query.filter(SaidaMaterial.data_movimento == data_busca)
            except:
                flash('Formato de data inválido. Use dd/mm/aaaa', 'error')
        elif filtro == 'responsavel':
            query = query.filter(usuario_responsavel.nome.ilike(f'%{busca}%'))
        elif filtro == 'solicitante':
            query = query.filter(usuario_solicitante.nome.ilike(f'%{busca}%'))
        elif filtro == 'setor':
            query = query.filter(UnidadeLocal.descricao.ilike(f'%{busca}%'))
        elif filtro == 'id':
            if busca.isdigit():
                query = query.filter(SaidaMaterial.id == int(busca))
    
    if status:
        query = query.filter(SaidaMaterial.status == status)
    if data_inicio:
        query = query.filter(SaidaMaterial.data_movimento >= parse_date(data_inicio))
    if data_fim:
        query = query.filter(SaidaMaterial.data_movimento <= parse_date(data_fim))
    if solicitante_id:
        query = query.filter(SaidaMaterial.solicitante_id == solicitante_id)
    if item_id:
        query = query.join(SaidaItem).filter(SaidaItem.item_id == item_id)
    
    # Estatísticas
    hoje = date.today()
    primeiro_dia_mes = date(hoje.year, hoje.month, 1)
    
    total_saidas = SaidaMaterial.query.filter_by(estornada=False).count()
    saidas_mes = SaidaMaterial.query.filter(
        SaidaMaterial.data_movimento >= primeiro_dia_mes,
        SaidaMaterial.estornada == False
    ).count()
    solicitantes_unicos = db.session.query(SaidaMaterial.solicitante_id).distinct().count()
    saidas_estornadas = SaidaMaterial.query.filter_by(estornada=True).count()
    
    saidas = query.order_by(desc(SaidaMaterial.data_movimento)).paginate(page=page, per_page=per_page)
    
    # Dados para filtros
    solicitantes = Usuario.query.order_by(Usuario.nome).all()
    itens = Item.query.order_by(Item.nome).all()
    
    return render_template('lista_saida.html', 
                         saidas=saidas,
                         solicitantes=solicitantes,
                         itens=itens,
                         data_inicio=data_inicio,
                         data_fim=data_fim,
                         solicitante_id=solicitante_id,
                         item_id=item_id,
                         status=status,
                         total_saidas=total_saidas,
                         saidas_mes=saidas_mes,
                         solicitantes_unicos=solicitantes_unicos,
                         saidas_estornadas=saidas_estornadas,
                         filtro=filtro,
                         busca=busca)

@saida_bp.route('/saida/nova', methods=['GET', 'POST'])
@login_required
def nova_saida():
    if request.method == 'POST':
        try:
            data_movimento = datetime.strptime(request.form['data_movimento'], '%Y-%m-%d').date()
            solicitante_id = request.form['solicitante_id']
            observacao = request.form.get('observacao', '')
            
            # Validações básicas
            if data_movimento > date.today():
                raise ValueError("Data de movimento não pode ser futura")
            
            if not solicitante_id:
                raise ValueError("Solicitante é obrigatório")
            
            # Criar nova saída
            saida = SaidaMaterial(
                data_movimento=data_movimento,
                solicitante_id=solicitante_id,
                usuario_id=current_user.id,
                observacao=observacao,
                numero_documento=gerar_numero_documento()
            )
            db.session.add(saida)
            db.session.flush()  # Obter o ID da saída

            # Processar itens
            itens = request.form.getlist('item_id[]')
            quantidades = request.form.getlist('quantidade[]')
            valores_unitarios = request.form.getlist('valor_unitario[]')

            if not itens:
                raise ValueError("É necessário incluir pelo menos um item")

            total_saida = Decimal('0.0')

            for item_id, qtd, valor_unit in zip(itens, quantidades, valores_unitarios):
                if item_id and qtd and valor_unit:
                    item = Item.query.get(item_id)
                    if not item:
                        raise ValueError(f"Item {item_id} não encontrado")

                    quantidade = int(qtd)
                    valor_unitario = Decimal(valor_unit.replace(',', '.'))

                    if quantidade <= 0:
                        raise ValueError("Quantidade deve ser maior que zero")

                    if item.estoque_atual < quantidade:
                        raise ValueError(f"Estoque insuficiente para o item {item.nome}")

                    # Validar valor unitário
                    if valor_unitario <= 0:
                        raise ValueError(f"Valor unitário inválido para o item {item.nome}")

                    if valor_unitario > item.valor_unitario * Decimal('1.1'):
                        raise ValueError(f"Valor unitário muito alto para o item {item.nome}")

                    saida_item = SaidaItem(
                        saida_id=saida.id,
                        item_id=item_id,
                        quantidade=quantidade,
                        valor_unitario=valor_unitario
                    )
                    db.session.add(saida_item)

                    # Atualizar estoque
                    item.estoque_atual -= quantidade
                    item.saldo_financeiro -= (quantidade * valor_unitario)
                    
                    # Atualizar total da saída
                    total_saida += (quantidade * valor_unitario)

                    # Verificar estoque mínimo
                    if item.estoque_atual <= item.estoque_minimo:
                        flash(f'Atenção: Item {item.nome} atingiu o estoque mínimo!', 'warning')

            db.session.commit()
            flash('Saída registrada com sucesso! Valor total: ' + format_currency(total_saida), 'success')
            return redirect(url_for('saida_bp.listar_saidas'))

        except ValueError as e:
            db.session.rollback()
            flash(str(e), 'error')
            return redirect(url_for('saida_bp.nova_saida'))
        except Exception as e:
            db.session.rollback()
            flash('Erro ao registrar saída: ' + str(e), 'error')
            return redirect(url_for('saida_bp.nova_saida'))

    # GET - Renderizar formulário
    itens = Item.query.filter(Item.estoque_atual > 0).order_by(Item.nome).all()
    itens_formatados = [{
        'id': item.id,
        'nome': item.nome,
        'valor_unitario': float(item.valor_unitario),
        'estoque_atual': float(item.estoque_atual),
        'unidade': item.unidade
    } for item in itens]
    
    solicitantes = Usuario.query.order_by(Usuario.nome).all()
    unidades = UnidadeLocal.query.all()
    return render_template('nova_saida.html', 
                         itens=itens_formatados, 
                         solicitantes=solicitantes,
                         unidades=unidades,
                         numero_documento=gerar_numero_documento(),
                         data_atual=date.today().strftime('%Y-%m-%d'))

@saida_bp.route('/saida/estornar/<int:saida_id>', methods=['POST'])
@login_required
def estornar_saida(saida_id):
    try:
        saida = SaidaMaterial.query.get_or_404(saida_id)
        
        if saida.estornada:
            return jsonify({'success': False, 'message': 'Esta saída já foi estornada'}), 400

        if saida.status != 'EFETIVADA':
            return jsonify({'success': False, 'message': 'Apenas saídas efetivadas podem ser estornadas'}), 400

        # Validar data do estorno
        if (date.today() - saida.data_movimento) > timedelta(days=30):
            return jsonify({
                'success': False, 
                'message': 'Não é possível estornar saídas com mais de 30 dias'
            }), 400

        # Estornar cada item
        for saida_item in saida.itens:
            item = saida_item.item
            item.estoque_atual += saida_item.quantidade
            item.saldo_financeiro += (saida_item.quantidade * saida_item.valor_unitario)

        saida.estornada = True
        
        # Se a saída está vinculada a uma requisição, também atualizar a requisição
        requisicao = RequisicaoMaterial.query.filter_by(saida_id=saida.id).first()
        if requisicao:
            requisicao.status = 'ESTORNADA'
            if requisicao.tarefa:
                requisicao.tarefa.status = 'Cancelada'
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Saída estornada com sucesso'})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@saida_bp.route('/api/item/<int:item_id>')
@login_required
def get_item_info(item_id):
    item = Item.query.get_or_404(item_id)
    return jsonify({
        'id': item.id,
        'nome': item.nome,
        'valor_unitario': float(item.valor_unitario),
        'estoque_atual': float(item.estoque_atual),
        'unidade': item.unidade
    })

@saida_bp.route('/saida/imprimir/<int:saida_id>')
@login_required
def imprimir_saida(saida_id):
    saida = SaidaMaterial.query.get_or_404(saida_id)
    return render_template('imprimir_saida.html', 
                         saida=saida,
                         format_currency=format_currency)

@saida_bp.route('/saida/detalhes/<int:saida_id>')
@login_required
def detalhes_saida(saida_id):
    saida = SaidaMaterial.query.get_or_404(saida_id)
    return render_template('detalhes_saida.html', 
                         saida=saida,
                         format_currency=format_currency)

@saida_bp.route('/saidas/relatorio')
@login_required
def relatorio_saidas():
    # Parâmetros do relatório
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    solicitante_id = request.args.get('solicitante_id', type=int)
    item_id = request.args.get('item_id', type=int)
    formato = request.args.get('formato', 'html')
    
    # Query base
    query = db.session.query(
        SaidaMaterial.data_movimento,
        Usuario.nome.label('solicitante'),
        Item.nome.label('item'),
        SaidaItem.quantidade,
        SaidaItem.valor_unitario,
        (SaidaItem.quantidade * SaidaItem.valor_unitario).label('valor_total')
    ).join(
        SaidaItem, Usuario, Item
    ).filter(
        SaidaMaterial.estornada == False
    )
    
    # Aplicar filtros
    if data_inicio:
        query = query.filter(SaidaMaterial.data_movimento >= parse_date(data_inicio))
    if data_fim:
        query = query.filter(SaidaMaterial.data_movimento <= parse_date(data_fim))
    if solicitante_id:
        query = query.filter(SaidaMaterial.solicitante_id == solicitante_id)
    if item_id:
        query = query.filter(SaidaItem.item_id == item_id)
    
    # Ordenação
    query = query.order_by(SaidaMaterial.data_movimento.desc())
    
    # Executar query
    resultados = query.all()
    
    if formato == 'excel':
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Saídas"
        
        # Cabeçalho
        headers = ['Data', 'Solicitante', 'Item', 'Quantidade', 'Valor Unitário', 'Valor Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Dados
        for row, dado in enumerate(resultados, 2):
            ws.cell(row=row, column=1, value=dado.data_movimento.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=dado.solicitante)
            ws.cell(row=row, column=3, value=dado.item)
            ws.cell(row=row, column=4, value=dado.quantidade)
            ws.cell(row=row, column=5, value=float(dado.valor_unitario))
            ws.cell(row=row, column=6, value=float(dado.valor_total))
        
        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        
        # Salvar arquivo
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'relatorio_saidas_{date.today()}.xlsx'
        )
        
    elif formato == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Cabeçalho
        writer.writerow(['Data', 'Solicitante', 'Item', 'Quantidade', 'Valor Unitário', 'Valor Total'])
        
        # Dados
        for dado in resultados:
            writer.writerow([
                dado.data_movimento.strftime('%d/%m/%Y'),
                dado.solicitante,
                dado.item,
                dado.quantidade,
                float(dado.valor_unitario),
                float(dado.valor_total)
            ])
        
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'relatorio_saidas_{date.today()}.csv'
        )
    
    # Formato HTML (default)
    return render_template(
        'relatorio.html',
        resultados=resultados,
        data_inicio=data_inicio,
        data_fim=data_fim,
        solicitante_id=solicitante_id,
        item_id=item_id,
        format_currency=format_currency
    )

@saida_bp.route('/saidas/dashboard')
@login_required
def dashboard():
    # Período padrão: último mês
    data_fim = date.today()
    data_inicio = data_fim - timedelta(days=30)
    
    # Top 5 itens mais saídos
    top_itens = db.session.query(
        Item.nome,
        func.sum(SaidaItem.quantidade).label('total_quantidade'),
        func.sum(SaidaItem.quantidade * SaidaItem.valor_unitario).label('total_valor')
    ).join(
        SaidaItem, SaidaMaterial
    ).filter(
        SaidaMaterial.data_movimento.between(data_inicio, data_fim),
        SaidaMaterial.estornada == False
    ).group_by(
        Item.id, Item.nome
    ).order_by(
        func.sum(SaidaItem.quantidade).desc()
    ).limit(5).all()
    
    # Total de saídas por dia
    saidas_por_dia = db.session.query(
        SaidaMaterial.data_movimento,
        func.count(SaidaMaterial.id).label('total_saidas'),
        func.sum(SaidaItem.quantidade * SaidaItem.valor_unitario).label('total_valor')
    ).join(
        SaidaItem
    ).filter(
        SaidaMaterial.data_movimento.between(data_inicio, data_fim),
        SaidaMaterial.estornada == False
    ).group_by(
        SaidaMaterial.data_movimento
    ).order_by(
        SaidaMaterial.data_movimento
    ).all()
    
    # Itens com estoque baixo
    itens_criticos = Item.query.filter(
        Item.estoque_atual <= Item.estoque_minimo
    ).order_by(
        (Item.estoque_atual / Item.estoque_minimo)
    ).limit(5).all()
    
    return render_template(
        'dashboard.html',
        top_itens=top_itens,
        saidas_por_dia=saidas_por_dia,
        itens_criticos=itens_criticos,
        format_currency=format_currency
    )

@saida_bp.route('/saida/requisicao/<int:saida_id>')
@login_required
def requisicao_saida(saida_id):
    saida = SaidaMaterial.query.get_or_404(saida_id)
    return render_template('requisicao_saida.html', 
                         saida=saida,
                         format_currency=format_currency)

@saida_bp.route('/saida/cancelar/<int:saida_id>', methods=['POST'])
@login_required
def cancelar_saida(saida_id):
    try:
        saida = SaidaMaterial.query.get_or_404(saida_id)
        
        if saida.status != 'PENDENTE':
            return jsonify({'success': False, 'message': 'Esta saída não está pendente'}), 400

        # Atualizar status da saída
        saida.status = 'CANCELADA'
        
        # Se a saída está vinculada a uma requisição, também atualizar a requisição
        requisicao = RequisicaoMaterial.query.filter_by(saida_id=saida.id).first()
        if requisicao:
            requisicao.status = 'CANCELADA'
            if requisicao.tarefa:
                requisicao.tarefa.status = 'Cancelada'
        
        db.session.commit()
        flash('Saída cancelada com sucesso!', 'success')
        return redirect(url_for('saida_bp.listar_saidas'))
    
    except Exception as e:
        db.session.rollback()
        flash('Erro ao cancelar saída: ' + str(e), 'error')
        return redirect(url_for('saida_bp.listar_saidas'))

